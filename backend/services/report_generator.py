import csv
import datetime
from io import StringIO
import io
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sqlmodel import Session, select
from models import JournalEntry
from constants.emotions import NEGATIVE_EMOTIONS

# ─────────────────────────────────────────────────────────────────────────────
# Range config
# ─────────────────────────────────────────────────────────────────────────────
RANGE_DAYS = {
    "weekly":  7,
    "monthly": 30,
    "yearly":  365,
}

# ─────────────────────────────────────────────────────────────────────────────
# EMOTION COLOR SCIENCE
#
# Each emotion is mapped to 3 values:
#   bg    — pastel background for the row/cell (light, readable)
#   badge — saturated color for count pills and accent borders
#   font  — dark text color from the same hue family (for contrast on bg)
#
# Psychological rationale:
#   Warm ambers/yellows → joy, amusement, excitement  (energy, brightness)
#   Greens              → gratitude, relief, caring    (growth, calm, safety)
#   Pinks/magentas      → love, desire, embarrassment  (connection, warmth)
#   Aquas/teals         → admiration, curiosity        (clarity, openness)
#   Sky blues           → pride, realization           (sky, freedom, insight)
#   Deep purples        → fear, nervousness            (dread, unease)
#   Mid blues           → sadness                     (cold, withdrawn)
#   Dark navies/slates  → grief, disappointment        (heavy, muted)
#   Warm oranges/reds   → anger, annoyance, disgust    (heat, aggression)
#   Indigo/violet       → remorse, disapproval         (regret, moral weight)
#   Neutral grays       → neutral, confusion           (baseline)
# ─────────────────────────────────────────────────────────────────────────────

EMOTION_PALETTE = {
    "joy":            {"bg": "FFF3C4", "badge": "F59E0B", "font": "78350F"},
    "amusement":      {"bg": "FEF9C3", "badge": "EAB308", "font": "713F12"},
    "excitement":     {"bg": "FFEDD5", "badge": "FB923C", "font": "7C2D12"},
    "gratitude":      {"bg": "DCFCE7", "badge": "22C55E", "font": "14532D"},
    "love":           {"bg": "FCE7F3", "badge": "EC4899", "font": "831843"},
    "admiration":     {"bg": "CCFBF1", "badge": "14B8A6", "font": "134E4A"},
    "pride":          {"bg": "E0F2FE", "badge": "0EA5E9", "font": "0C4A6E"},
    "optimism":       {"bg": "ECFCCB", "badge": "84CC16", "font": "365314"},
    "relief":         {"bg": "D1FAE5", "badge": "10B981", "font": "064E3B"},
    "caring":         {"bg": "DCFCE7", "badge": "4ADE80", "font": "166534"},
    "approval":       {"bg": "D1FAE5", "badge": "34D399", "font": "064E3B"},
    "curiosity":      {"bg": "E0F2FE", "badge": "38BDF8", "font": "0C4A6E"},
    "realization":    {"bg": "DBEAFE", "badge": "60A5FA", "font": "1E3A8A"},
    "surprise":       {"bg": "CFFAFE", "badge": "22D3EE", "font": "164E63"},
    "desire":         {"bg": "FCE7F3", "badge": "F472B6", "font": "831843"},
    "anger":          {"bg": "FEE2E2", "badge": "EF4444", "font": "7F1D1D"},
    "annoyance":      {"bg": "FFEDD5", "badge": "F97316", "font": "7C2D12"},
    "disgust":        {"bg": "ECFCCB", "badge": "65A30D", "font": "365314"},
    "fear":           {"bg": "EDE9FE", "badge": "7C3AED", "font": "2E1065"},
    "nervousness":    {"bg": "F3E8FF", "badge": "A855F7", "font": "3B0764"},
    "sadness":        {"bg": "DBEAFE", "badge": "3B82F6", "font": "1E3A8A"},
    "grief":          {"bg": "E2E8F0", "badge": "475569", "font": "0F172A"},
    "disappointment": {"bg": "F1F5F9", "badge": "64748B", "font": "1E293B"},
    "remorse":        {"bg": "EEF2FF", "badge": "6366F1", "font": "1E1B4B"},
    "embarrassment":  {"bg": "FDF2F8", "badge": "C026D3", "font": "4A044E"},
    "disapproval":    {"bg": "FEE2E2", "badge": "DC2626", "font": "7F1D1D"},
    "confusion":      {"bg": "F8FAFC", "badge": "94A3B8", "font": "1E293B"},
    "neutral":        {"bg": "F9FAFB", "badge": "9CA3AF", "font": "1F2937"},
}

def _get_palette(emotion: str) -> dict:
    return EMOTION_PALETTE.get(emotion.strip().lower(), EMOTION_PALETTE["neutral"])

def _blend_multi_emotion(emotions: list) -> dict:
    """Blend top-2 emotion palettes for multi-emotion rows."""
    if not emotions:
        return EMOTION_PALETTE["neutral"]
    if len(emotions) == 1:
        return _get_palette(emotions[0])
    p1, p2 = _get_palette(emotions[0]), _get_palette(emotions[1])
    def avg(h1, h2):
        r = (int(h1[0:2],16)+int(h2[0:2],16))//2
        g = (int(h1[2:4],16)+int(h2[2:4],16))//2
        b = (int(h1[4:6],16)+int(h2[4:6],16))//2
        return f"{r:02X}{g:02X}{b:02X}"
    return {"bg": avg(p1["bg"],p2["bg"]), "badge": avg(p1["badge"],p2["badge"]), "font": p1["font"]}

def _lighten(hex_color: str, factor: float) -> str:
    """Blend a hex color toward white. factor=0 original, factor=1 white."""
    try:
        r,g,b = int(hex_color[0:2],16), int(hex_color[2:4],16), int(hex_color[4:6],16)
        return f"{int(r+(255-r)*factor):02X}{int(g+(255-g)*factor):02X}{int(b+(255-b)*factor):02X}"
    except Exception:
        return "FFFFFF"

def _fill(h):    return PatternFill(start_color=h, end_color=h, fill_type="solid")
def _font(h, bold=False, size=10, name="Calibri", italic=False):
    return Font(color=h, bold=bold, size=size, name=name, italic=italic)
def _border(color="E2E8F0", left_color=None, left_style="thin"):
    s = Side(style="thin", color=color)
    l = Side(style=left_style, color=left_color or color)
    return Border(left=l, right=s, top=s, bottom=s)
def _align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

# ─────────────────────────────────────────────────────────────────────────────
# Emotion-aware suggestions
# ─────────────────────────────────────────────────────────────────────────────
EMOTION_SUGGESTIONS = {
    "anger":          "Box breathing (4s in · 4s hold · 4s out · 4s hold) + a brisk walk",
    "annoyance":      "Step away for 10 min — a short break resets the frustration loop",
    "fear":           "5-4-3-2-1 grounding: name 5 things you can see right now",
    "nervousness":    "Slow extended exhale (6s out). Progressive muscle relaxation helps",
    "sadness":        "Behavioral activation — start with one small achievable task",
    "grief":          "Allow the feeling. Journaling or calling a trusted person helps",
    "disappointment": "Reframe: what did this teach you? Small wins rebuild momentum",
    "remorse":        "Acknowledge it, repair if possible, then practice self-compassion",
    "embarrassment":  "Everyone has these moments. Write it out to defuse the feeling",
    "disapproval":    "Channel into constructive feedback rather than internal rumination",
    "disgust":        "Distance from the trigger. Grounding breath sequence helps",
    "confusion":      "Brain dump everything onto paper — clarity follows externalization",
    "joy":            "Savor it consciously. Share with someone or write it down",
    "amusement":      "Lean in — laughter is recovery. Seek more of what caused this",
    "excitement":     "Channel into your current project while the energy is high",
    "gratitude":      "Write 3 specific things you're grateful for — specificity deepens it",
    "love":           "Express it — a message, a call, or one small kind act today",
    "admiration":     "Study what you admire. It often reveals your own deeper values",
    "pride":          "Note what led here. Repeat those habits deliberately next time",
    "optimism":       "Plan one concrete next step toward your goal while energized",
    "relief":         "Rest fully — the nervous system genuinely needs recovery time",
    "caring":         "You give a lot. Make sure you're also receiving care today",
    "approval":       "Positive signal. Reflect on what's working and do more of it",
    "curiosity":      "Follow this thread — read, search, or ask someone who knows",
    "realization":    "Write the insight down before it fades. Act on one part today",
    "surprise":       "Take a breath. Give yourself time to process before reacting",
    "desire":         "Check: does this align with your deeper values before pursuing?",
    "neutral":        "Stable baseline — a great time for planning or focused deep work",
}

def _get_suggestion(emotions: list) -> str:
    for e in emotions:
        if e.strip() in EMOTION_SUGGESTIONS:
            return EMOTION_SUGGESTIONS[e.strip()]
    return EMOTION_SUGGESTIONS["neutral"]


# ─────────────────────────────────────────────────────────────────────────────
# CSV REPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_csv_report(user_id, range_type, engine):
    days       = RANGE_DAYS.get(range_type, 30)
    end_date   = datetime.date.today()
    start_date = end_date - datetime.timedelta(days=days)

    with Session(engine) as session:
        entries = session.exec(
            select(JournalEntry)
            .where(JournalEntry.user_id == user_id)
            .where(JournalEntry.date >= start_date)
            .order_by(JournalEntry.date)
        ).all()

    output = StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)

    # Header block
    writer.writerow(["Intelligent Emotion Recognition and Reminder System — Emotional Wellness Report"])
    writer.writerow([f"Period: {range_type.capitalize()}  |  {start_date.strftime('%d %b %Y')} → {end_date.strftime('%d %b %Y')}"])
    writer.writerow([f"Total entries: {len(entries)}  |  Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}"])
    writer.writerow([])

    # Summary stats
    all_emotions = []
    for e in entries:
        all_emotions.extend([x.strip() for x in e.emotion_primary.split(",")])

    if all_emotions:
        counts  = Counter(all_emotions)
        top3    = counts.most_common(3)
        neg_cnt = sum(1 for e in entries if any(x.strip() in NEGATIVE_EMOTIONS for x in e.emotion_primary.split(",")))
        pos_cnt = len(entries) - neg_cnt
        writer.writerow(["=== SUMMARY ==="])
        writer.writerow(["Top emotions:", "  |  ".join(f"{e.capitalize()} ({c})" for e, c in top3)])
        writer.writerow(["Positive entries:", pos_cnt, "  Negative entries:", neg_cnt])
        writer.writerow([])

    # Column headers
    writer.writerow(["Date", "Mood", "Emotion(s) Detected", "Confidence", "Journal Entry", "Wellness Suggestion"])

    # Data rows
    for entry in entries:
        emotions  = [e.strip() for e in entry.emotion_primary.split(",")]
        is_neg    = any(e in NEGATIVE_EMOTIONS for e in emotions)
        mood      = "Negative" if is_neg else "Positive"
        top_score = ""
        if entry.emotion_scores and emotions:
            score = entry.emotion_scores.get(emotions[0])
            if score is not None:
                top_score = f"{round(score * 100)}%"
        writer.writerow([
            f'="{entry.date.strftime("%Y-%m-%d")}"',
            mood,
            ", ".join(e.capitalize() for e in emotions),
            top_score,
            entry.input_sentence or "[No content]",
            _get_suggestion(emotions),
        ])

    writer.writerow([])
    writer.writerow([f"Intelligent Emotion Recognition and Reminder System  |  Confidential  |  {datetime.datetime.now().strftime('%d %b %Y')}"])

    return output.getvalue(), start_date, end_date


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT — 2 sheets
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel_report(user_id, range_type, engine):
    days       = RANGE_DAYS.get(range_type, 30)
    end_date   = datetime.date.today()
    start_date = end_date - datetime.timedelta(days=days)

    with Session(engine) as session:
        entries = session.exec(
            select(JournalEntry)
            .where(JournalEntry.user_id == user_id)
            .where(JournalEntry.date >= start_date)
            .order_by(JournalEntry.date)
        ).all()

    wb = openpyxl.Workbook()
    _build_journal_sheet(wb.active, entries, range_type, start_date, end_date)
    ws2 = wb.create_sheet("Emotion Summary")
    _build_summary_sheet(ws2, entries, range_type, start_date, end_date)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), start_date, end_date


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Journal Log
# ─────────────────────────────────────────────────────────────────────────────

def _build_journal_sheet(ws, entries, range_type, start_date, end_date):
    ws.title = "Journal Log"
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False

    # Row heights
    for r, h in {1:38, 2:20, 3:16, 4:8, 5:26}.items():
        ws.row_dimensions[r].height = h

    # Column widths: A=Date B=Mood C=Emotions D=Confidence E=Journal F=Suggestion
    for col, w in zip("ABCDEF", [14, 13, 26, 13, 52, 52]):
        ws.column_dimensions[col].width = w

    # ── Row 1: Brand bar ──────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "  \U0001F9E0   Intelligent Emotion Recognition and Reminder System  —  Emotional Wellness Report"
    c.fill      = _fill("0F172A")
    c.font      = _font("F1F5F9", bold=True, size=15)
    c.alignment = _align("left", "center", indent=1)

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value     = f"  {range_type.capitalize()} Report   ·   {len(entries)} entries   ·   {start_date.strftime('%d %b %Y')} → {end_date.strftime('%d %b %Y')}"
    c.fill      = _fill("1E293B")
    c.font      = _font("64748B", size=9)
    c.alignment = _align("left", "center", indent=1)

    # ── Row 3: Generated timestamp ────────────────────────────────────────────
    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value     = f"  Generated {datetime.datetime.now().strftime('%d %b %Y at %H:%M')}"
    c.fill      = _fill("334155")
    c.font      = _font("94A3B8", size=8, italic=True)
    c.alignment = _align("left", "center", indent=1)

    # ── Row 4: Thin spacer ────────────────────────────────────────────────────
    for col in range(1, 7):
        ws.cell(row=4, column=col).fill = _fill("F8FAFC")

    # ── Row 5: Column headers ─────────────────────────────────────────────────
    headers = ["DATE", "MOOD", "EMOTION(S)", "CONFIDENCE", "JOURNAL ENTRY", "WELLNESS SUGGESTION"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=ci)
        c.value     = h
        c.fill      = _fill("4338CA")
        c.font      = _font("E0E7FF", bold=True, size=9)
        c.alignment = _align("center", "center")
        c.border    = Border(bottom=Side(style="medium", color="3730A3"))

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, entry in enumerate(entries, start=6):
        ws.row_dimensions[ri].height = 44

        emotions = [e.strip() for e in entry.emotion_primary.split(",")]
        is_neg   = any(e in NEGATIVE_EMOTIONS for e in emotions)
        palette  = _blend_multi_emotion(emotions)

        top_score = ""
        if entry.emotion_scores and emotions:
            score = entry.emotion_scores.get(emotions[0])
            if score is not None:
                top_score = f"{round(score * 100)}%"

        # Column A — Date
        c = ws.cell(row=ri, column=1)
        c.value     = entry.date.strftime("%d %b %Y")
        c.fill      = _fill("F8FAFC")
        c.font      = _font("1E293B", bold=True, size=10)
        c.alignment = _align("center", "center")
        c.border    = _border(left_color=palette["badge"], left_style="medium")

        # Column B — Mood indicator
        c = ws.cell(row=ri, column=2)
        mood_bg    = _lighten("FEE2E2", 0.3) if is_neg else _lighten("DCFCE7", 0.3)
        mood_color = "B91C1C" if is_neg else "15803D"
        mood_text  = "▼  Negative" if is_neg else "▲  Positive"
        c.value     = mood_text
        c.fill      = _fill(mood_bg)
        c.font      = _font(mood_color, bold=True, size=9)
        c.alignment = _align("center", "center")
        c.border    = _border()

        # Column C — Emotions (the star — full emotion color science)
        c = ws.cell(row=ri, column=3)
        c.value     = "  ·  ".join(e.capitalize() for e in emotions)
        c.fill      = _fill(palette["bg"])
        c.font      = _font(palette["font"], bold=True, size=9)
        c.alignment = _align("center", "center", wrap=True)
        c.border    = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="medium", color=palette["badge"]),
            bottom=Side(style="medium", color=palette["badge"]),
        )

        # Column D — Confidence (badge color fill for visual punch)
        c = ws.cell(row=ri, column=4)
        c.value     = top_score
        c.fill      = _fill(palette["badge"])
        c.font      = _font("FFFFFF", bold=True, size=11)
        c.alignment = _align("center", "center")
        c.border    = _border()

        # Column E — Journal entry (very light tint of emotion color)
        c = ws.cell(row=ri, column=5)
        c.value     = entry.input_sentence or "[No content]"
        c.fill      = _fill(_lighten(palette["bg"], 0.55))
        c.font      = _font("334155", size=9, italic=True)
        c.alignment = _align("left", "top", wrap=True, indent=1)
        c.border    = _border()

        # Column F — Suggestion (white bg, palette font color)
        c = ws.cell(row=ri, column=6)
        c.value     = _get_suggestion(emotions)
        c.fill      = _fill("FFFFFF")
        c.font      = _font(palette["font"], size=9)
        c.alignment = _align("left", "top", wrap=True, indent=1)
        c.border    = _border()

    # ── Footer ────────────────────────────────────────────────────────────────
    fr = len(entries) + 7
    ws.merge_cells(f"A{fr}:F{fr}")
    c = ws[f"A{fr}"]
    c.value     = f"  Intelligent Emotion Recognition and Reminder System   ·   Confidential wellness data   ·   {datetime.datetime.now().strftime('%d %b %Y')}"
    c.fill      = _fill("0F172A")
    c.font      = _font("475569", size=8, italic=True)
    c.alignment = _align("center", "center")
    ws.row_dimensions[fr].height = 18


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Emotion Frequency Summary
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, entries, range_type, start_date, end_date):
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    for r, h in {1:36, 2:18, 3:8, 4:24}.items():
        ws.row_dimensions[r].height = h
    for col, w in zip("ABCDE", [22, 14, 14, 32, 22]):
        ws.column_dimensions[col].width = w

    # Brand bar
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = "  \U0001F9E0   Emotion Frequency Summary"
    c.fill      = _fill("0F172A")
    c.font      = _font("F1F5F9", bold=True, size=14)
    c.alignment = _align("left", "center", indent=1)

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value     = f"  {range_type.capitalize()}  ·  {start_date.strftime('%d %b %Y')} → {end_date.strftime('%d %b %Y')}  ·  {len(entries)} entries"
    c.fill      = _fill("1E293B")
    c.font      = _font("64748B", size=9)
    c.alignment = _align("left", "center", indent=1)

    for col in range(1, 6):
        ws.cell(row=3, column=col).fill = _fill("F8FAFC")

    # Headers
    for ci, h in enumerate(["EMOTION", "COUNT", "% OF ENTRIES", "FREQUENCY", "CATEGORY"], 1):
        c = ws.cell(row=4, column=ci)
        c.value     = h
        c.fill      = _fill("312E81")
        c.font      = _font("E0E7FF", bold=True, size=9)
        c.alignment = _align("center", "center")
        c.border    = Border(bottom=Side(style="medium", color="4338CA"))

    # Tally
    all_emotions = []
    for e in entries:
        all_emotions.extend([x.strip() for x in e.emotion_primary.split(",")])
    counts = Counter(all_emotions)
    total  = max(len(entries), 1)
    sorted_emotions = sorted(counts.items(), key=lambda x: -x[1])

    for ri, (emotion, count) in enumerate(sorted_emotions, start=5):
        ws.row_dimensions[ri].height = 22
        palette  = _get_palette(emotion)
        pct      = round((count / total) * 100, 1)
        bar_fill = "█" * min(int(pct / 3), 30)
        bar_empty= "░" * (30 - len(bar_fill))
        is_neg   = emotion in NEGATIVE_EMOTIONS

        row_vals = [emotion.capitalize(), count, f"{pct}%", bar_fill + bar_empty, "Negative" if is_neg else "Positive"]

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci)
            c.value = val

            if ci == 1:
                c.fill      = _fill(palette["bg"])
                c.font      = _font(palette["font"], bold=True, size=10)
                c.alignment = _align("left", "center", indent=1)
                c.border    = _border(left_color=palette["badge"], left_style="medium")

            elif ci == 2:
                c.fill      = _fill(palette["badge"])
                c.font      = _font("FFFFFF", bold=True, size=11)
                c.alignment = _align("center", "center")
                c.border    = _border()

            elif ci == 3:
                c.fill      = _fill(palette["bg"])
                c.font      = _font(palette["font"], size=10)
                c.alignment = _align("center", "center")
                c.border    = _border()

            elif ci == 4:
                c.fill      = _fill("F8FAFC")
                c.font      = Font(color=palette["badge"], size=8, name="Courier New")
                c.alignment = _align("left", "center")
                c.border    = _border()

            elif ci == 5:
                c.fill      = _fill("FEE2E2" if is_neg else "DCFCE7")
                c.font      = _font("991B1B" if is_neg else "166534", bold=True, size=9)
                c.alignment = _align("center", "center")
                c.border    = _border()

    # Footer
    fr = len(sorted_emotions) + 6
    ws.merge_cells(f"A{fr}:E{fr}")
    c = ws[f"A{fr}"]
    c.value     = f"  Intelligent Emotion Recognition and Reminder System   ·   {datetime.datetime.now().strftime('%d %b %Y')}"
    c.fill      = _fill("0F172A")
    c.font      = _font("475569", size=8, italic=True)
    c.alignment = _align("center", "center")
    ws.row_dimensions[fr].height = 16